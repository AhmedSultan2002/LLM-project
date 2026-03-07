"""
Data Preprocessing Module
=========================
Parses the NUST Bank product knowledge Excel workbook and the funds transfer
FAQ JSON file into a unified, cleaned document corpus for embedding.

Usage:
    python src/data_preprocessing.py
"""

import json
import os
import re
import sys

import openpyxl

# Allow running as script from project root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.settings import RAW_EXCEL_PATH, RAW_JSON_PATH, DATA_DIR, PROCESSED_DOCS_PATH


# ──────────────────────────────────────────────────────────────────────────────
# Text Cleaning
# ──────────────────────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    """Normalize unicode, whitespace, and strip a text string."""
    if not text or not isinstance(text, str):
        return ""
    # Replace smart quotes with standard ones
    text = text.replace("\u201c", '"').replace("\u201d", '"')
    text = text.replace("\u2018", "'").replace("\u2019", "'")
    # Replace right arrow with ->
    text = text.replace("\u2192", "->")
    # Collapse multiple whitespace / newlines into single space
    text = re.sub(r"\s+", " ", text).strip()
    return text


# ──────────────────────────────────────────────────────────────────────────────
# Excel Parser
# ──────────────────────────────────────────────────────────────────────────────

def parse_excel(filepath: str) -> list[dict]:
    """
    Parse all sheets of the NUST Bank Product Knowledge workbook.

    Each sheet follows a pattern:
      - Row 1: Product title (e.g. "NUST Savings Account")
      - Row 2: blank
      - Rows 3+: alternating question / answer pairs
        (question in one row, answer in the next)

    The data is primarily in columns A or B depending on the sheet.
    """
    wb = openpyxl.load_workbook(filepath)
    documents = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip empty sheets
        if ws.max_row is None or ws.max_row < 3:
            continue

        # Collect all non-empty text cells from each row (take first non-None value)
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            # Find the first non-None cell with actual text
            cell_text = None
            for cell in row:
                if cell is not None and str(cell).strip():
                    cell_text = str(cell).strip()
                    break
            rows_text.append(cell_text)

        # Extract product title from the first non-empty row
        product_title = None
        for text in rows_text:
            if text:
                product_title = clean_text(text)
                break

        if not product_title:
            continue

        # Extract Q&A pairs: look for question-then-answer patterns
        # Questions typically end with '?' or start with common question words
        i = 0
        while i < len(rows_text):
            text = rows_text[i]
            if text and _is_question(text):
                question = clean_text(text)
                # Look ahead for the answer (next non-empty, non-question row)
                answer_parts = []
                j = i + 1
                while j < len(rows_text):
                    next_text = rows_text[j]
                    if next_text is None:
                        j += 1
                        continue
                    if _is_question(next_text):
                        break  # Hit the next question
                    answer_parts.append(clean_text(next_text))
                    j += 1

                answer = " ".join(answer_parts).strip()
                if answer:
                    documents.append({
                        "question": question,
                        "answer": answer,
                        "product": product_title,
                        "source": f"excel:{sheet_name}",
                    })
                i = j  # Skip to where we left off
            else:
                i += 1

    return documents


def _is_question(text: str) -> bool:
    """Heuristic: decide if a row of text is a question."""
    text = text.strip()
    if text.endswith("?"):
        return True
    lower = text.lower()
    question_starters = [
        "what ", "how ", "can ", "is ", "does ", "do ", "will ",
        "are ", "which ", "who ", "when ", "where ", "why ",
        "for which", "in case",
    ]
    return any(lower.startswith(s) for s in question_starters)


# ──────────────────────────────────────────────────────────────────────────────
# JSON Parser
# ──────────────────────────────────────────────────────────────────────────────

def parse_json_faq(filepath: str) -> list[dict]:
    """Parse the funds transfer app features FAQ JSON file."""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    documents = []
    for category in data.get("categories", []):
        cat_name = category.get("category", "General")
        for qa in category.get("questions", []):
            question = clean_text(qa.get("question", ""))
            answer = clean_text(qa.get("answer", ""))
            if question and answer:
                documents.append({
                    "question": question,
                    "answer": answer,
                    "product": f"Mobile App — {cat_name}",
                    "source": "json:funds_transfer_faq",
                })
    return documents


# ──────────────────────────────────────────────────────────────────────────────
# Corpus Builder
# ──────────────────────────────────────────────────────────────────────────────

def build_corpus() -> list[dict]:
    """
    Parse all data sources and build a unified document corpus.

    Each document has:
        - question: str
        - answer: str
        - product: str
        - source: str
        - text: str  (combined chunk for embedding)
    """
    print("Parsing Excel workbook...")
    excel_docs = parse_excel(RAW_EXCEL_PATH)
    print(f"  → Extracted {len(excel_docs)} Q&A pairs from Excel")

    print("Parsing JSON FAQ...")
    json_docs = parse_json_faq(RAW_JSON_PATH)
    print(f"  → Extracted {len(json_docs)} Q&A pairs from JSON")

    # Combine
    all_docs = excel_docs + json_docs

    # Build the combined text chunk for each document
    for doc in all_docs:
        doc["text"] = (
            f"Product: {doc['product']}\n"
            f"Q: {doc['question']}\n"
            f"A: {doc['answer']}"
        )

    print(f"\nTotal corpus: {len(all_docs)} documents")
    return all_docs


def save_corpus(documents: list[dict], output_path: str) -> None:
    """Save the processed corpus to JSON."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(documents, f, indent=2, ensure_ascii=False)
    print(f"Saved to {output_path}")


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    corpus = build_corpus()
    save_corpus(corpus, PROCESSED_DOCS_PATH)

    # Print a few samples
    print("\n── Sample Documents ──")
    for doc in corpus[:3]:
        print(f"\n[{doc['source']}] {doc['product']}")
        print(f"  Q: {doc['question']}")
        print(f"  A: {doc['answer'][:120]}...")
